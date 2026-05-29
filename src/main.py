import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
from classifier import analyze_pdf
from parser_debit import parse_debit_pdf
from parser_credit import parse_credit_pdf
import threading

# ─── Palette ────────────────────────────────────────────────────────────────
BBVA_BLUE    = "#004481"
BBVA_NAVY    = "#001F5B"
BBVA_SKY     = "#1973B8"
ACCENT_CYAN  = "#00C4D4"
BG_DARK      = "#0A1628"
BG_CARD      = "#0F1F3D"
BG_FIELD     = "#152340"
TEXT_WHITE   = "#FFFFFF"
TEXT_MUTED   = "#8899BB"
SUCCESS_GRN  = "#00D68F"
ERROR_RED    = "#FF4B6E"
WARN_AMBER   = "#FFB830"
FONT_MAIN    = "Helvetica"


class LedgerFlowApp:
    def __init__(self, root):
        self.root = root
        self.root.title("LedgerFlow BBVA")
        self.root.geometry("640x520")
        self.root.resizable(False, False)
        self.root.configure(bg=BG_DARK)

        self.source_dir    = tk.StringVar()
        self.pdf_count     = 0
        self.processed     = 0
        self._build_ui()

    # ─── UI construction ───────────────────────────────────────────────────
    def _build_ui(self):
        # ── top stripe ────────────────────────────────────────────────────
        stripe = tk.Frame(self.root, bg=BBVA_BLUE, height=4)
        stripe.pack(fill=tk.X, side=tk.TOP)

        # ── main container ────────────────────────────────────────────────
        outer = tk.Frame(self.root, bg=BG_DARK)
        outer.pack(fill=tk.BOTH, expand=True, padx=30, pady=20)

        # ── hero header ───────────────────────────────────────────────────
        hdr = tk.Frame(outer, bg=BG_DARK)
        hdr.pack(fill=tk.X, pady=(0, 22))

        logo_lbl = tk.Label(hdr, text="LedgerFlow", font=(FONT_MAIN, 26, "bold"),
                            bg=BG_DARK, fg=TEXT_WHITE)
        logo_lbl.pack(side=tk.LEFT)

        tag_lbl = tk.Label(hdr, text="  BBVA", font=(FONT_MAIN, 26),
                           bg=BG_DARK, fg=BBVA_SKY)
        tag_lbl.pack(side=tk.LEFT)

        ver_lbl = tk.Label(hdr, text="v2.0", font=(FONT_MAIN, 10),
                           bg=BG_DARK, fg=TEXT_MUTED)
        ver_lbl.place_forget()  # just declare; not shown
        ver_lbl.pack(side=tk.RIGHT, pady=12)

        sub_lbl = tk.Label(outer,
                           text="Extractor automático de estados de cuenta BBVA · Débito & Crédito",
                           font=(FONT_MAIN, 10), bg=BG_DARK, fg=TEXT_MUTED)
        sub_lbl.pack(anchor=tk.W, pady=(0, 20))

        # ── folder card ───────────────────────────────────────────────────
        card = tk.Frame(outer, bg=BG_CARD, bd=0, relief=tk.FLAT)
        card.pack(fill=tk.X, pady=(0, 14))
        self._round_border(card)

        inner = tk.Frame(card, bg=BG_CARD)
        inner.pack(fill=tk.X, padx=16, pady=14)

        fold_lbl = tk.Label(inner, text="📁  Carpeta de PDFs",
                            font=(FONT_MAIN, 10, "bold"),
                            bg=BG_CARD, fg=TEXT_MUTED)
        fold_lbl.pack(anchor=tk.W, pady=(0, 6))

        entry_row = tk.Frame(inner, bg=BG_CARD)
        entry_row.pack(fill=tk.X)

        self.entry = tk.Entry(entry_row, textvariable=self.source_dir,
                              font=(FONT_MAIN, 11), bg=BG_FIELD, fg=TEXT_WHITE,
                              insertbackground=TEXT_WHITE, bd=0,
                              relief=tk.FLAT, highlightthickness=1,
                              highlightcolor=BBVA_SKY, highlightbackground=BG_CARD)
        self.entry.pack(side=tk.LEFT, fill=tk.X, expand=True,
                        ipady=8, padx=(0, 10))

        browse_btn = self._mk_btn(entry_row, "Buscar", self._browse, width=9,
                                  bg=BBVA_SKY, fg=TEXT_WHITE)
        browse_btn.pack(side=tk.RIGHT)

        # ── stats bar ─────────────────────────────────────────────────────
        stats_card = tk.Frame(outer, bg=BG_CARD)
        stats_card.pack(fill=tk.X, pady=(0, 14))

        self.stat_pdf  = self._stat_block(stats_card, "PDFs encontrados", "—")
        self.stat_deb  = self._stat_block(stats_card, "Cuentas débito",   "—")
        self.stat_cred = self._stat_block(stats_card, "Tarjetas crédito", "—")

        for f in stats_card.winfo_children():
            f.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=6, pady=8)

        # ── progress bar ─────────────────────────────────────────────────
        pb_frame = tk.Frame(outer, bg=BG_DARK)
        pb_frame.pack(fill=tk.X, pady=(0, 6))

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("BBVA.Horizontal.TProgressbar",
                        troughcolor=BG_CARD, background=ACCENT_CYAN,
                        darkcolor=ACCENT_CYAN, lightcolor=ACCENT_CYAN,
                        bordercolor=BG_CARD, thickness=6)
        self.pb = ttk.Progressbar(pb_frame, style="BBVA.Horizontal.TProgressbar",
                                  mode="determinate", length=580)
        self.pb.pack(fill=tk.X)

        # ── status label ─────────────────────────────────────────────────
        self.status_var = tk.StringVar(value="Selecciona una carpeta para comenzar")
        self.status_lbl = tk.Label(outer, textvariable=self.status_var,
                                   font=(FONT_MAIN, 10), bg=BG_DARK, fg=TEXT_MUTED)
        self.status_lbl.pack(anchor=tk.W, pady=(4, 14))

        # ── CTA button ───────────────────────────────────────────────────
        self.proc_btn = self._mk_btn(
            outer, "⚡   PROCESAR Y EXPORTAR A EXCEL",
            self._start_processing,
            bg=BBVA_BLUE, fg=TEXT_WHITE,
            font=(FONT_MAIN, 13, "bold"), ipady=14,
            state=tk.DISABLED
        )
        self.proc_btn.pack(fill=tk.X)

        # ── bottom stripe ─────────────────────────────────────────────────
        bot = tk.Frame(self.root, bg=BBVA_NAVY, height=3)
        bot.pack(fill=tk.X, side=tk.BOTTOM)

    # ─── helper widgets ────────────────────────────────────────────────────
    def _mk_btn(self, parent, text, command, bg=BBVA_BLUE, fg=TEXT_WHITE,
                font=None, width=None, ipady=8, state=tk.NORMAL):
        kw = dict(text=text, command=command, bg=bg, fg=fg,
                  activebackground=BBVA_SKY, activeforeground=TEXT_WHITE,
                  bd=0, relief=tk.FLAT, cursor="hand2", state=state)
        if font:  kw["font"] = font
        else:     kw["font"] = (FONT_MAIN, 11, "bold")
        if width: kw["width"] = width
        return tk.Button(parent, **kw)

    def _stat_block(self, parent, label, value):
        f = tk.Frame(parent, bg=BG_FIELD)
        tk.Label(f, text=label, font=(FONT_MAIN, 8), bg=BG_FIELD,
                 fg=TEXT_MUTED).pack(pady=(6, 0))
        v = tk.Label(f, text=value, font=(FONT_MAIN, 18, "bold"),
                     bg=BG_FIELD, fg=ACCENT_CYAN)
        v.pack(pady=(0, 6))
        return v   # return the value label so we can update it

    def _round_border(self, frame):
        """Simulate a subtle left accent border with a coloured strip."""
        accent = tk.Frame(frame, bg=BBVA_SKY, width=3)
        accent.pack(side=tk.LEFT, fill=tk.Y)

    # ─── events ────────────────────────────────────────────────────────────
    def _browse(self):
        folder = filedialog.askdirectory()
        if not folder:
            return
        self.source_dir.set(folder)
        pdfs = [f for f in os.listdir(folder) if f.lower().endswith('.pdf')]
        self.pdf_count = len(pdfs)
        self.stat_pdf.config(text=str(self.pdf_count))
        self.stat_deb.config(text="—")
        self.stat_cred.config(text="—")
        self.pb['value'] = 0
        if pdfs:
            self._set_status(f"{self.pdf_count} archivo(s) PDF encontrado(s) · Listo para procesar", ACCENT_CYAN)
            self.proc_btn.config(state=tk.NORMAL, bg=BBVA_BLUE)
        else:
            self._set_status("No se encontraron PDFs en esa carpeta", ERROR_RED)
            self.proc_btn.config(state=tk.DISABLED)

    def _set_status(self, msg, color=TEXT_MUTED):
        self.status_var.set(msg)
        self.status_lbl.config(fg=color)

    def _start_processing(self):
        """Launch processing in a background thread to keep UI responsive."""
        self.proc_btn.config(state=tk.DISABLED)
        self._set_status("Procesando…", WARN_AMBER)
        t = threading.Thread(target=self._process_statements, daemon=True)
        t.start()

    # ─── core processing ───────────────────────────────────────────────────
    def _process_statements(self):
        folder = self.source_dir.get()
        if not folder or not os.path.exists(folder):
            self._set_status("La carpeta seleccionada no existe", ERROR_RED)
            self.proc_btn.config(state=tk.NORMAL)
            return

        pdf_files = [os.path.join(folder, f)
                     for f in os.listdir(folder) if f.lower().endswith('.pdf')]
        if not pdf_files:
            self._set_status("Sin PDFs en la carpeta", ERROR_RED)
            self.proc_btn.config(state=tk.NORMAL)
            return

        debit_master  = {}
        credit_master = {}
        debit_count = credit_count = 0
        errors = []
        total = len(pdf_files)

        for idx, pdf_path in enumerate(pdf_files):
            fname = os.path.basename(pdf_path)
            self._set_status(f"Procesando {fname}…", WARN_AMBER)

            try:
                doc_type, period = analyze_pdf(pdf_path)
            except Exception as e:
                errors.append(f"{fname}: {e}")
                self._update_progress(idx + 1, total)
                continue

            try:
                if doc_type == "DEBIT":
                    rows = parse_debit_pdf(pdf_path)
                    if rows:
                        debit_master.setdefault(period, []).extend(rows)
                        debit_count += 1
                elif doc_type == "CREDIT":
                    rows = parse_credit_pdf(pdf_path)
                    if rows:
                        credit_master.setdefault(period, []).extend(rows)
                        credit_count += 1
                else:
                    errors.append(f"{fname}: formato no reconocido")
            except Exception as e:
                errors.append(f"{fname}: {e}")

            self._update_progress(idx + 1, total)

        # ── update stat blocks ────────────────────────────────────────────
        self.root.after(0, lambda: self.stat_deb.config(text=str(debit_count)))
        self.root.after(0, lambda: self.stat_cred.config(text=str(credit_count)))

        if debit_count + credit_count == 0:
            self._set_status("Ningún PDF correspondió a un formato BBVA compatible", ERROR_RED)
            self.root.after(0, lambda: self.proc_btn.config(state=tk.NORMAL, bg=BBVA_BLUE))
            return

        out_files = []

        # ── Débito Excel ──────────────────────────────────────────────────
        if debit_master:
            path = os.path.join(folder, "LedgerFlow_Debito.xlsx")
            with pd.ExcelWriter(path, engine='openpyxl') as writer:
                for month in sorted(debit_master.keys()):
                    df = pd.DataFrame(debit_master[month])
                    self._style_sheet(df, writer, month)
            out_files.append("LedgerFlow_Debito.xlsx")

        # ── Crédito Excel ─────────────────────────────────────────────────
        if credit_master:
            path = os.path.join(folder, "LedgerFlow_Credito.xlsx")
            with pd.ExcelWriter(path, engine='openpyxl') as writer:
                for month in sorted(credit_master.keys()):
                    df = pd.DataFrame(credit_master[month])
                    self._style_sheet(df, writer, month)
            out_files.append("LedgerFlow_Credito.xlsx")

        # ── Done ──────────────────────────────────────────────────────────
        msg_parts = [f"✔  {f}" for f in out_files]
        if errors:
            msg_parts.append(f"\n⚠  {len(errors)} archivo(s) con errores")
        self._set_status("¡Exportación completa!  →  " + "  |  ".join(out_files), SUCCESS_GRN)

        self.root.after(0, lambda: messagebox.showinfo(
            "LedgerFlow — Proceso Terminado",
            "Archivos generados en la carpeta de origen:\n\n"
            + "\n".join(f"  •  {f}" for f in out_files)
            + (("\n\nAdvertencias:\n" + "\n".join(f"  - {e}" for e in errors[:5])) if errors else "")
        ))
        self.root.after(0, lambda: self.proc_btn.config(state=tk.NORMAL, bg=BBVA_BLUE))

    # ─── progress helper ───────────────────────────────────────────────────
    def _update_progress(self, done, total):
        pct = int(done / total * 100)
        self.root.after(0, lambda: self.pb.config(value=pct))

    # ─── Excel styling ─────────────────────────────────────────────────────
    def _style_sheet(self, df, writer, sheet_name):
        from openpyxl.styles import (PatternFill, Font, Alignment,
                                     Border, Side)
        from openpyxl.utils import get_column_letter

        df.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]

        HEADER_FILL  = PatternFill("solid", fgColor="004481")
        ALT_FILL     = PatternFill("solid", fgColor="EFF4FB")
        HEADER_FONT  = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
        BODY_FONT    = Font(name="Calibri", size=10)
        MONEY_FONT   = Font(name="Calibri", size=10)
        CENTER       = Alignment(horizontal="center", vertical="center")
        thin         = Side(border_style="thin", color="C5D1E8")
        border       = Border(left=thin, right=thin, bottom=thin, top=thin)

        # Header row
        for cell in ws[1]:
            cell.fill       = HEADER_FILL
            cell.font       = HEADER_FONT
            cell.alignment  = CENTER
            cell.border     = border

        # Data rows
        money_cols = {col_idx for col_idx, col in enumerate(df.columns, 1)
                      if df[col].dtype in ['float64', 'int64']}

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
            for cell in row:
                cell.fill      = fill
                cell.border    = border
                cell.alignment = CENTER if cell.column in money_cols \
                                       else Alignment(vertical="center", wrap_text=True)
                cell.font      = MONEY_FONT if cell.column in money_cols else BODY_FONT
                if cell.column in money_cols and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'

        # Auto-width
        for col_idx, col in enumerate(df.columns, 1):
            letter  = get_column_letter(col_idx)
            max_len = max(len(str(col)), df[col].astype(str).str.len().max())
            ws.column_dimensions[letter].width = min(max_len + 4, 42)

        ws.freeze_panes = "A2"


if __name__ == "__main__":
    root = tk.Tk()
    root.configure(bg=BG_DARK)
    app = LedgerFlowApp(root)
    root.mainloop()
